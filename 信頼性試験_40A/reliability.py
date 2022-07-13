import os
import pathlib
from pathlib import Path
import glob
import re
import cx_Oracle as oracle
import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.preprocessing import PolynomialFeatures
import openpyxl

##定数の定義
cutoff_voltageC = 4.225
cutoff_voltageD = 2.8
terminal_resistance = 0.05
# 一次近似判定閾値
term1 = 0.999
# 二次近似判定閾値
term2 = 0.998


# フィルター列名
# JT放電容量
fil_dischargeV = "Cyc"
# Mode
fil_Mode = "Mode"
# StpTime[sec]
fil_StpTime = "StpTime[sec]"
# DatF
fil_DatF = "DatF"
# Cycle
fil_Cycle = "Cycle"
# StepNo
fil_StepNo = "StepNo."

# 選択列名
# 放電容量
col_discharV = "放電量[Ah]"
# 電流
col_current = "Curr[A]"
# 到達電圧
col_vol = "Volt[V]"
# 開始電圧
col_initV = "InitVolt[V]"
# 開始温度
col_initTemp = "InitTemp1[℃]"
# 到達温度
col_temp = "Temp1[℃]"


# 計算条件
first_order = "直線近似"
second_order = "曲線近似"
two_point_order = "二次近似"

##関数の定義
#フォルダ名から使用するファイル名のリストを作成する関数
def make_file_path_list(mode=1, ch_no=None, file_type=None, csv_list=None):
    
    '''
    mode: 1 -> 必要な全てのファイルリストを返す 
          2 -> 任意のファイルのパスを返す
    file_type: 1 -> 放電容量
               2 -> 充電抵抗
               3 -> CycleEnd
               4 -> 放電抵抗25
               5 -> 放電抵抗-10
    '''
    
    if mode == 1:

        # カレントディレクトリを取得
        cwd = os.getcwd()
        
        # データフォルダ名を取得
        target_folders = []
        for f in glob.glob(cwd + "\\*"):
            folder_name = os.path.split(f)[1]
            if re.findall("^\d{8}", folder_name):
                target_folders.append(folder_name)
                
        
        # 対象ディレクトリのパスを作成(list)
        path_list = [cwd + "\\" + target_folder + "\\*\\**" for target_folder in target_folders]
        
        
        # 対象ディレクトリから全てのファイル名を取得(list)
        data_folders = []

        for path in path_list:
            data_folders.append(glob.glob(path))

            
        # 各データフォルダから必要なファイルを抽出し、各チャネル号機毎に5つのファイルを含んだリストを作成する
        # カレントフォルダにあるデータフォルダの個数分の長さのリストができる
        csv_lists = []
        for data_folder in data_folders:

            file_list=[]

            # ファイルリストから必要なファイルのみをフィルタリング
            for file in data_folder:
                file_list += [file]

            #各チャンネル毎で必要なcsvファイルのパスを取得し、リスト化
            CH_NO = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14']
            csv_list = []
            for ch in CH_NO:
                a = ch + '_CapacityCyc'
                b = ch + '_Count000005Ptn005'
                c = ch + '_CycleEnd'
                d = ch + '_Count000009Ptn009'
                e = ch + '_Count000014Ptn014'


                f = [file for file in file_list if ('Grp' in file) and ( a in file) ]
                f += [file for file in file_list if ('Grp' in file) and ( b in file) ]
                f += [file for file in file_list if ('Grp' in file) and ( c in file) ]
                f += [file for file in file_list if ('Grp' in file) and ( d in file) ]
                f += [file for file in file_list if ('Grp' in file) and ( e in file) ]

                csv_list.append(f)

            # 各データフォルダの最初のチャネル番号を取得してcsv_listにセットで格納
            for idx, ch in enumerate(csv_list, start=1):
                if ch:
                    csv_list.insert(0, idx)
                    break        

            csv_lists.append(csv_list)    
        return csv_lists
    
    elif mode == 2:
        return csv_list[ch_no][file_type-1]
    
    
#CSVファイルのヘッダ行を取得する関数
def get_header(csv_file=None):
    with open(csv_file) as f:
        for row, text in enumerate(f, start=0):
            text = text.rstrip()
            if text == r'Data_Start':
                dst_row = row
                break
        else:
            raise ValueError('Not Found')
        
    return dst_row + 1

#CSVファイルを結合して1つのデータフレームを作成する関数
def make_dataframe(csv_ptn=None, csv_cycleEnd=None, filter_mode=1):
    
    """
    filter_mode: 1=充電抵抗(5sec)_at25℃_SOC50%, 
                 2=放電抵抗(10sec)_at25℃_SOC80%, 
                 3=放電抵抗(10sec)_at-10℃_SOC30%
    """
    if filter_mode == 1:
        df_ptn = pd.read_csv(csv_ptn, header=get_header(csv_ptn), encoding='shift-jis')
        df_ptn = df_ptn[(df_ptn[fil_Mode] == "1:定電流充電(CC)") & (df_ptn[fil_StpTime] == 5)]
        df_ptn = df_ptn.loc[:,[col_current, col_vol, col_temp]].reset_index(drop=True)
        df_CycleEnd = pd.read_csv(csv_cycleEnd, header=get_header(csv_cycleEnd), encoding='shift-jis')
        df_CycleEnd = df_CycleEnd[(df_CycleEnd[fil_Cycle] == 5) & \
                    (df_CycleEnd[fil_Mode] == "001:定電流充電(CC)") & (df_CycleEnd[fil_DatF] == "77:時間")]
        df_CycleEnd = df_CycleEnd.loc[:,[col_initV, col_initTemp]].reset_index(drop=True)
    else:
        df_ptn = pd.read_csv(csv_ptn, header=get_header(csv_ptn), encoding='shift-jis')
        df_ptn = df_ptn[(df_ptn[fil_StpTime] == 10) & \
                        ((df_ptn[fil_StepNo] == 1) | \
                        (df_ptn[fil_StepNo] == 9) | \
                        (df_ptn[fil_StepNo] == 17) | \
                        (df_ptn[fil_StepNo] == 25) | \
                        (df_ptn[fil_StepNo] == 33) | \
                        (df_ptn[fil_StepNo] == 41) | \
                        (df_ptn[fil_StepNo] == 49) | \
                        (df_ptn[fil_StepNo] == 57))]
        df_ptn = df_ptn.loc[:,[col_current, col_vol, col_temp]].reset_index(drop=True)
        df_ptn[col_current] = df_ptn[col_current].apply(lambda x: -x)
        df_CycleEnd = pd.read_csv(csv_cycleEnd, header=get_header(csv_cycleEnd), encoding='shift-jis')
        if filter_mode == 2:
            df_CycleEnd = df_CycleEnd[(df_CycleEnd[fil_Cycle] == 9) & \
                        (df_CycleEnd[fil_Mode] == "006:定電流放電(DC)") & (df_CycleEnd[fil_DatF] == "77:時間")]
        elif filter_mode == 3:
            df_CycleEnd = df_CycleEnd[(df_CycleEnd[fil_Cycle] == 14) & \
                        (df_CycleEnd[fil_Mode] == "006:定電流放電(DC)") & (df_CycleEnd[fil_DatF] == "77:時間")]
        df_CycleEnd = df_CycleEnd.loc[:,[col_initV, col_initTemp]].reset_index(drop=True)
        
        
    df = pd.concat([df_ptn, df_CycleEnd], axis=1)
    
    return df

#放電容量を取得する関数
def get_discharge_capacity(csv_dc=None):
    df = pd.read_csv(csv_dc, header=get_header(csv_dc), encoding='shift-jis')
    
    # JT放電容量のフィルター条件: Cyc = 2
    df_Cyc2 = df[df["Cyc"] == 2]
    
    discharge_capacity = df_Cyc2[col_discharV].iloc[-1]
    
    return discharge_capacity

#カット電圧到達点の有無判断と、あった場合の到達点数を取得する関数
def get_index(dataF=None, cutoff_voltage=1):
    """
    cutoff_volage: 1=充電抵抗, 2=放電抵抗
    """
    # どのカット電圧値を使用するかを切り分け
    if cutoff_voltage == 1:
        dataF_over = dataF[dataF[col_vol] >= cutoff_voltageC]
    else:
        dataF_over = dataF[dataF[col_vol] <= cutoff_voltageD]
    
    # カット電圧到達点がある場合
    if len(dataF_over):
        before_index = dataF.index.stop - len(dataF_over) - 1
        after_index = dataF.index.stop - len(dataF_over)
        return True, len(dataF_over)
    else:
        return False, 0
    
#内部抵抗および最終抵抗を算出する関数
def calc_final_resistance(initV=None, cutoffV=None, current=None):
    
    # 内部抵抗の算出
    internal_resistance = abs(initV - cutoffV) / current * 1000
    # 最終抵抗値 = 内部抵抗 - 端子抵抗
    final_resistance = internal_resistance - terminal_resistance
    
    return final_resistance

#カット電圧未達の場合の抵抗算出の関数
# とりあえずDataFrameを引数に設定
def calc_resistance_with_effective_value(dataF=None):
    
    # point8の開始電圧を内部抵抗算出の開始電圧として使用
    initialV = dataF[col_initV].loc[dataF.shape[0] - 1]
    # point8の到達電圧を内部抵抗算出の到達電圧として使用
    ultimateV = dataF[col_vol].loc[dataF.shape[0] - 1]
    # point8の電流値を内部抵抗算出の電流値として使用
    current = dataF[col_current].loc[dataF.shape[0] - 1]
    
    return calc_final_resistance(initV=initialV, cutoffV=ultimateV, current=current)

#一次近似の可不可を判断する関数
def judge_first_order_approximation(dataF=None, plotNum=None, output_corr=False):
    """
    plotNum: カット電圧到達数
    output_results: False -> 出力しない, True -> 出力する
    """
    
    # データフレームの範囲を決定
    dataF = dataF.iloc[:8 - plotNum + 1]
    # DataFrame.corr()**2で相関係数を算出
    df_correlation = dataF.corr()**2
    correlation = df_correlation.iat[0, 1]
    
    if output_corr == True:
        print(correlation)
    # 相関係数 > 0.999なら一次近似可と判断
    return True if correlation > 0.999 else False

#一次近似の処理をする関数
def calc_first_order_approximation(dataF=None, plotNum=None, cutoff_voltage=1, output_results=False):
    """
    plotNum: カット電圧到達数
    cutoff_volage: 1=充電抵抗, 2=放電抵抗
    output_results: False -> 出力しない, True -> 出力する
    """    
    
    # データフレームの範囲を決定
    dataF = dataF.iloc[:8 - plotNum + 1]
    
    # カット到達前後の開始電圧の平均を内部抵抗算出の為の開始電圧とする
    # 結合したデータフレームでは3列目が開始電圧
    beforeV = dataF.iat[-plotNum, 3]
    afterV = dataF.iat[-plotNum + 1, 3]
    initialV = (beforeV + afterV) / 2

    # データフレームの各列を２次元に変換
    # 到達電圧:説明変数　電流:目的変数
    x = dataF[[col_vol]]
    y = dataF[[col_current]]
    
    # scikit-learnを使って一次近似(係数：切片を算出)
    model = LinearRegression()
    model.fit(x, y)

    coefficient, intercept = model.coef_[0][0], model.intercept_[0]
    
    # 回帰式を使ってカット電圧値における電流値を算出
    # 充電抵抗/放電抵抗によるカット電圧値切り替え
    if cutoff_voltage == 1:
        cutoffV = cutoff_voltageC
    else:
        cutoffV = cutoff_voltageD
    
    current = coefficient * cutoffV + intercept 

    return calc_final_resistance(initV=initialV, cutoffV=cutoffV, current=current)

#二次近似の可不可および計算をする関数
def judgeCalc_second_order_approximation(dataF=None, plotNum=None, cutoff_voltage=1, output_corr=False):
    """
    plotNum: カット電圧到達数
    cutoff_volage: 1=充電抵抗, 2=放電抵抗
    output_results: False -> 出力しない, True -> 出力する
    """
    
    # データフレームの範囲を決定
    dataF = dataF.iloc[:8 - plotNum + 1]
    
    # カット到達前後の開始電圧の平均を内部抵抗算出の為の開始電圧とする
    # 結合したデータフレームでは3列目が開始電圧
    beforeV = dataF.iat[-plotNum, 3]
    afterV = dataF.iat[-plotNum + 1, 3]
    initialV = (beforeV + afterV) / 2

    # データフレームの各列を２次元に変換
    # 到達電圧:説明変数　電流:目的変数
    x = dataF[[col_vol]]
    y = dataF[[col_current]]

    # 二次近似式の決定係数を算出
    # 説明変数のデータを２次式用に加工
    polynomial_features = PolynomialFeatures(degree=2)
    x_poly = polynomial_features.fit_transform(x)

    # y = b0 + b1x + b2x^2 の b0～b2 を算出
    model = LinearRegression()
    model.fit(x_poly, y)
    y_pred = model.predict(x_poly)

    # 決定係数
    r2 = r2_score(y, y_pred)

    # 決定係数 >= 0.998 の場合二次近似を行う
    if r2 >= 0.998:
        # 各係数および定数項を算出
        coefficient1, coefficient2, const = model.coef_[0][1], model.coef_[0][2], model.intercept_[0]

        # 充電抵抗/放電抵抗によるカット電圧値切り替え
        if cutoff_voltage == 1:
            cutoffV = cutoff_voltageC
        else:
            cutoffV = cutoff_voltageD
            
        # 回帰式を用いてカット電圧値における電流値を求める
        current = coefficient2 * (cutoffV ** 2) + coefficient1 * cutoffV + const

        # V = IRで内部抵抗を算出
        return True, calc_final_resistance(initV=initialV, cutoffV=cutoffV, current=current)
    # 二次近似不可の場合はFalseを返す
    else:
        return False, 0

#二点近似の処理
def calc_two_points_approximation(dataF=None, plotNum=None, cutoff_voltage=1, output_corr=False):
    
    # データフレームの範囲を決定
    dataF = dataF.iloc[:8 - plotNum + 1]
    
    # カット到達前後の開始電圧の平均を内部抵抗算出の為の開始電圧とする
    # 結合したデータフレームでは3列目が開始電圧
    beforeV = dataF.iat[-plotNum, 3]
    afterV = dataF.iat[-plotNum + 1, 3]
    initialV = (beforeV + afterV) / 2

    
    # 到達前後の到達電圧のリストを作成
    beforeV = dataF.iat[-plotNum, 1]
    afterV = dataF.iat[-plotNum + 1, 1]
    v = [beforeV, afterV]

    # 到達前後の電流値のリストを作成
    beforeC = dataF.iat[-plotNum, 0]
    afterC = dataF.iat[-plotNum + 1, 0]
    c = [beforeC, afterC]

    # 到達前後の２点を使って単回帰
    coefficient, intercept = np.polyfit(v, c, 1)
    
    # 回帰式を使ってカット電圧値における電流値を算出
    # 充電抵抗/放電抵抗によるカット電圧値切り替え
    if cutoff_voltage == 1:
        cutoffV = cutoff_voltageC
    else:
        cutoffV = cutoff_voltageD
    
    current = coefficient * cutoffV + intercept 

    return calc_final_resistance(initV=initialV, cutoffV=cutoffV, current=current)

#一連の計算処理をまとめた関数（対応チャンネル数1）
def calc_results(csv_dischageV=None, csv_chargingR=None, csv_CycleEnd=None, csv_dischargeR25=None, csv_dischargeR10=None):
    # JT放電容量を求める
    discharge_capacity = get_discharge_capacity(csv_dischageV)

    # 充電抵抗・放電抵抗25・放電抵抗-10についてそれぞれデータフレームを作成
    dfC = make_dataframe(csv_chargingR, csv_CycleEnd, filter_mode=1)
    dfD25 = make_dataframe(csv_dischargeR25, csv_CycleEnd, filter_mode=2)
    dfD10 = make_dataframe(csv_dischargeR10, csv_CycleEnd, filter_mode=3)

    # カット電圧到達の有無を判断(充電抵抗・放電抵抗25・放電抵抗-10)
    reachedC, pointsC = get_index(dfC, cutoff_voltage=1)
    reachedD25, pointsD25 = get_index(dfD25, cutoff_voltage=2)
    reachedD10, pointsD10 = get_index(dfD10, cutoff_voltage=2)

    # 各項目のインデックスと結果を対応させたdictを作成
    # [到達有無, df, 到達点数, カットオフ電圧値]
    items_dict = {0: [reachedC, dfC, pointsC, 1], 1: [reachedD25, dfD25, pointsD25, 2], 2: [reachedD10, dfD10, pointsD10, 2]}

    results_list = []
    
    # 最終計算条件: カット電圧未達/直線/曲線/二次近似
    condition = ""
    results_list.append(discharge_capacity)

    for key, value_list in items_dict.items():

        # 充電抵抗
        if value_list[0]:
            # 一次近似可不可の判断
            first_order_ok = judge_first_order_approximation(dataF=value_list[1], plotNum=value_list[2], output_corr=False)

            # 一次近似可の場合の処理
            if first_order_ok:
                resistance = calc_first_order_approximation(dataF=value_list[1], plotNum=value_list[2], cutoff_voltage=value_list[3], output_results=False)
                condition += "直線, "
            # 一次近似不可の場合の処理
            else:
                # 二次近似の可の場合
                second_order_ok, resistance = judgeCalc_second_order_approximation(dataF=value_list[1], plotNum=value_list[2], cutoff_voltage=value_list[3], output_corr=False)
                condition += "曲線, "
                if not second_order_ok:
                    # 二次近似不可の場合は二点近似
                    resistance = calc_two_points_approximation(dataF=value_list[1], plotNum=value_list[2], cutoff_voltage=value_list[3], output_corr=False)
                    condition += "二次近似, "
        else:
            resistance = calc_resistance_with_effective_value(dataF=value_list[1])
            condition += "カット電圧未達, "

        results_list.append(resistance)
        
    results_list.append(condition[:-2])
    
    return results_list

#抜取セルデータ用フレーム作成関数
def make_sample_df(excel_path=None, csv_path=None):
    
    # Excelファイルの読み込み
    df_sample = pd.read_excel(excel_path, header=2)
    
    # 検査ロットをファイル内のデータから抽出
    with open(csv_path) as f:
        for row, text in enumerate(f, start=0):
            text = text.rstrip()
            if "ロットNo." in text:
                lot_no = text
                break
        else:
            raise ValueError('Not Found')

    # ロットNoをあるだけ取り出す
    res = re.findall("\w{0,3}\d{2,3}", lot_no)

    lot_list = []
    for idx, lot in enumerate(res):
        if idx == 0:
            prefix = lot[:2]
        elif len(lot) == 3:
            lot = prefix + lot
        # csvファイルに'_'が含まれていた場合に取り除く (2022/06/08追加)
        if '_' in lot:
            lot = lot.replace('_', '', 1)
        lot_list.append(lot)

    df_list = []
    for lot in lot_list:
        if len(df_sample[df_sample["検査LOT"] == lot]):
            df_list.append(df_sample[df_sample["検査LOT"] == lot])
    
    df_sample = pd.concat(df_list, ignore_index=True)
    
    # for lot in lot_list:
    #     print(lot)
    #     if len(df_sample[df_sample["検査LOT"] == lot]):
    #         df_sample += df_sample[df_sample["検査LOT"] == lot]
            
    # サイクル終了予定日を全行にコピー(全行共通)
    # "サイクル終了予定日"がExcelの18行目
    df_sample["サイクル終了予定日"] = df_sample.iat[0, 18]
    
    #NaN -> 空文字に変換(NaNのままだとDBにインサートできない)
    df_sample.fillna("", inplace=True)
    
    # 各セルの寸法幅が規格内かどうかを判定
    df_sample["JudgeWid"] = df_sample["幅寸法mm"].apply(check_cell_width)
    
    return df_sample

#演算前テーブル作成用関数
# チャネル毎のDF -> 全チャネル分を結合
def make_bofore_calc_df(csv_list=None, sample_df=None):
    
    # データが空のチャネルをリストから除く
    csv_list = [csv for csv in csv_list if csv]
    
    # チャネル毎のDFのリストを作成
    df_lists = []
    # チャネル毎の放電容量を格納するための空のdictを準備
    dc_dict = {}
    for ch_list in csv_list:
        # 充電抵抗・放電抵抗25・放電抵抗-10についてそれぞれデータフレームを作成
        dfC = make_dataframe(ch_list[1], ch_list[2], filter_mode=1)
        dfD25 = make_dataframe(ch_list[3], ch_list[2], filter_mode=2)
        dfD10 = make_dataframe(ch_list[4], ch_list[2], filter_mode=3)

        df_list = [dfC, dfD25, dfD10]
        df_lists.append(df_list)
        
        # ファイル名からチャネル番号を取り出す
        ch = ""
        ch = re.findall("\d\d_CapacityCyc", ch_list[0])[0][:2]
        if ch[0] == '0':
            ch = ch[1]
        
        # 各チャネル毎の放電容量を抽出しdictに格納
        # ch_list[0] -> xx_CapacityCyc-000k.csv という名前のファイルを取ってきている
        dc_dict[ch] = get_discharge_capacity(ch_list[0])
        
        
    # 抜取検査DFの必要な列を追加
    result_list = []
    for ch, df_list in enumerate(df_lists):
        for idx, df in enumerate(df_list):

            if idx == 0:
                df["PtnNo"] = 5
            elif idx == 1:
                df["PtnNo"] = 9
            else:
                df["PtnNo"] = 14
            df["KLot"] = sample_df.values[ch][0]
            df["LineNo"] = sample_df.values[ch][1]
            df["MachineNo"] = sample_df.values[ch][5][-1]
            df["ChNo"] = re.findall("\d{1,2}", sample_df.values[ch][6])[0]
            df["CellId"] = sample_df.values[ch][2]
            result_list.append(df)
    
    # ここまで作成したDFを結合
    final_df = pd.concat(result_list)
    
    
    # 日付・担当者を抽出(フォルダ内にあるチャネル共通なので最後に一括で挿入)
    with open(csv_list[0][0]) as f:
        folder_name = ""
        person_name = ""
        for row, text in enumerate(f, start=0):
            text = text.rstrip()
            if "ﾃﾞｰﾀ保存ﾌｫﾙﾀﾞ" in text:
                folder_name = text
            if "担当者" in text:
                person_name = text

            if (folder_name and person_name):
                break

    d = re.findall("\d{8}", folder_name)[0]
    person = re.findall('"(.*)"', person_name)[0]
    
    
    # 日付・放電容量・担当者を追加
    final_df["Date"] = d
    final_df["Pic"] = person
    final_df["DischargeCap[Ah]"] = final_df["ChNo"].apply(lambda ch: dc_dict[ch])
    
    # インデックス張り直し
    final_df = final_df.reset_index(drop=True)
    
    # DBの列の並びに整頓
    final_df = final_df.reindex(columns=["Date", "KLot", "LineNo", \
                                     "MachineNo", "ChNo", "DischargeCap[Ah]", "PtnNo", \
                                     "Curr[A]", "Volt[V]", "Temp1[℃]", "InitVolt[V]",\
                                     "InitTemp1[℃]", "Pic", "CellId"])
    
    return final_df

#演算後データフレーム作成関数
def make_after_calc_df(csv_list=None, sample_df=None):
    
    # 日付を格納する変数を準備
    d = ""
    
    # チャネル毎の演算結果をリストに格納
    result_calc = []
    for idx, ch_list in enumerate(csv_list, 1):
        if ch_list:
            result_calc.append([str(idx), calc_results(*ch_list)])
    
            # 日付・担当者を抽出
            if not d:
                with open(make_file_path_list(2, idx, 1, csv_list)) as f:
                    folder_name = ""
                    for row, text in enumerate(f, start=0):
                        text = text.rstrip()
                        if "ﾃﾞｰﾀ保存ﾌｫﾙﾀﾞ" in text:
                            folder_name = text
                d = re.findall("\d{8}", folder_name)[0]


    # 演算結果のリストにチャネルを追加
    result_list = []
    for ch, l in result_calc:
        l.append(ch)
        result_list.append(l)

    # 作成したリストをDFに変換
    df_after = pd.DataFrame(result_list, columns=["DischargeCap[Ah]", "ChargeRes(5sec25℃Soc50%)",
                                                "DischargeRes(10sec25℃SOC80%)", "DischargeRes(10sec-10℃Soc30%)", 
                                                "Approximation", "Ch"])

    # サンプルデータフレームから必要な列を抽出
    df_sample = sample_df.iloc[:, [0, 1, 2, 5, 6]].copy()
    
    # サンプルデータフレームの列名を変更
    df_sample.rename(columns={'検査LOT':'Klot','検査ライン':'LineNo', 'セルID':'CellId', '投入サイクル機No':'MachineNo'}, inplace=True)
    
    # チャネル・サイクル機Noの数字部分のみ取り出す
    df_sample["Ch"] = df_sample["Ch"].apply(lambda s: s.lstrip("ch"))
    df_sample["MachineNo"] = df_sample["MachineNo"].apply(lambda s: s.lstrip("#"))
    
    # 演算結果DFと抜取DFを結合
    df_after = pd.merge(df_after, df_sample, how="inner", on="Ch")
    
    # 日付列の追加
    df_after["Date"] = d
    
    # 各項目の結果判定
    df_after["judge_dc"] = df_after["DischargeCap[Ah]"].apply(judge_discharge_capacity)
    df_after["judge_cr"] = df_after["ChargeRes(5sec25℃Soc50%)"].apply(judge_charging_resistance)
    df_after["judge_dr25"] = df_after["DischargeRes(10sec25℃SOC80%)"].apply(judge_discharing_resistance25)
    df_after["judge_dr10"] = df_after["DischargeRes(10sec-10℃Soc30%)"].apply(judge_discharing_resistance10)
    df_after["judge"] = ""
    
    # 総合結果判定
    for idx, data in df_after.iterrows():
        judge_list = [judge for judge in data[11:15]]
        if any([elem == 'NG' for elem in judge_list]):
            # 15列目が'judge'列
            df_after.iloc[idx, 15] = 'NG'
        else:
            df_after.iloc[idx, 15] = 'OK'
                          
    # DB用に列の並びを変更
    df_after = df_after.reindex(columns=["Date", "Klot", "LineNo", \
                                         "MachineNo", "Ch", "Approximation", "DischargeCap[Ah]", \
                                         "ChargeRes(5sec25℃Soc50%)","DischargeRes(10sec25℃SOC80%)", \
                                         "DischargeRes(10sec-10℃Soc30%)", "CellId", \
                                         "judge_dc", "judge_cr", "judge_dr25", "judge_dr10", "judge"])
    
    return df_after

##MESサーバーへ接続
#CX oracleでMESサーバーへ接続（接続情報以外は定型）
class conn_MES_LWR:
    def __init__(self, host= "10.60.28.21", port="1521", service="psh1dbv",
                       scheme="tabuser",username="tabuser",password="tab123"):        
        self.host = host
        self.port = port
        self.service  = service

        self.scheme   = scheme
        self.username = username
        self.password = password
    
    def __enter__(self):
        
        # tns:Oracleが命名したDB接続用インターフェース技術の名前
        
        # インターフェイスオブジェクトの作成
        self.tns  = oracle.makedsn(self.host, self.port, service_name=self.service) if self.host else None
        # 接続を確立
        self.conn = oracle.connect(self.username, self.password, self.tns) if self.tns else None
        # カーソルの取得
        self.curs = self.conn.cursor() if self.conn else None
        return self

    def __exit__(self, exception_type, exception_value, traceback):
        if self.curs is not None: self.curs.close()
        if self.conn is not None: self.conn.close()
        
#SQL
# 抜取テーブル用SQL
insert_sample = 'INSERT INTO "TABUSER"."ReliabilityTest(Smp)40A" ("KLot", "KlineNo", "CellId", "CellWid[mm]", "JigNo", "CycleNo", "ChNo", "RecDate", "InputDate", "TestPtn", "CurrentLineTerminalRes(+)[mΩ]", "CurrentLineTerminalRes(-)[mΩ]", "CurrentLineTotalRes[mΩ]", "VoltageLineTerminalRes(+)[mΩ]", "VoltageLineTerminalRes(-)[mΩ]", "VoltageLineTotalRes[mΩ]", "JudgeRes", "Note", "CycleEndDate", "JudgeWid") VALUES (:1,:2,:3,:4,:5,:6,:7,:8,:9,:10,:11,:12,:13,:14,:15,:16,:17,:18,:19,:20)'
# 演算前テーブル用SQL
insert_before = 'INSERT INTO "TABUSER"."ReliabilityTest(BefCal)40A" ("TestDate", "Klot", "LineNo", "MachineNo", "ChNo", "DischargeCap[Ah]", "PtnNo", "Curr[A]", "Volt[V]", "Temp1[℃]", "InitVolt[V]", "InitTemp1[℃]", "Pic", "CellId") VALUES (:1,:2,:3,:4,:5,:6,:7,:8,:9,:10,:11,:12,:13,:14)'
# 演算後テーブル用SQL
insert_after = 'INSERT INTO "TABUSER"."ReliabilityTest(AftCal)40A " ("Date", "Klot", "LineNo", "MachineNo", "ChNo", "Approximation", "DischargeCap[Ah]", "ChargeRes(5sec25℃Soc50%)", "DischargeRes(10sec25℃SOC80%)", "DischargeRes(10sec-10℃Soc30%)", "CellId", "DischargeCapResult", "ChargeRes(5sec25℃Soc50%)Reulst", "DischargeRes(10sec25℃SOC80%)Result", "DischargeRes(10sec-10℃Soc30%)Result", "FinalResult") VALUES (:1,:2,:3,:4,:5,:6,:7,:8,:9,:10,:11,:12,:13,:14,:15,:16)'

#セル幅寸法の判定用関数
def check_cell_width(width=None):
    if not (307.8 < width < 308.5):
        return "NG"
    else: 
        return "OK"
    
#各項目の規格内判定用関数
# 結果判定用の関数
def judge_discharge_capacity(dc=None):
    if 193.0 <= dc <= 209.0:
        return "OK"
    else:
        return "NG"

def judge_charging_resistance(cr=None):
    if 0.34 <= cr <= 0.52:
        return "OK"
    else:
        return "NG"
    
def judge_discharing_resistance25(dr=None):
    if 0.39 <= dr <= 0.59:
        return "OK"
    else:
        return "NG"
    
def judge_discharing_resistance10(dr=None):
    if 1.25 <= dr <= 1.87:
        return "OK"
    else:
        return "NG"

    
#DBにデータをアップする関数
def upload_data(insert_sql=None, df=None):
    
    # DBに格納できるようにデータフレームを二次元配列に変換
    rows = [list(x) for x in df.values]
    
    with conn_MES_LWR() as mesdb:
        # executemany()で複数行のデータを一括でインサート
        mesdb.curs.executemany(insert_sql, rows)
        mesdb.conn.commit()

#2次元配列のデータを一括でExcelに書き込む関数
# ※pd.to_execl()は上書きができないためopenpyxlを使用
def write_list_2d(sheet, l_2d, start_row, start_col):
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                       column=start_col + x,
                       value=l_2d[y][x])

#Excelに結果を出力する関数
def output_to_excel(df_after=None, df_sample=None, csv_list=None, ch_no=None):
    
    # カレントディレクトリを取得
    cwd = os.getcwd()
    
    # 結果出力先のExcelファイル名を取得
    for f in glob.glob(cwd + "\\*"):
        file_name = os.path.split(f)[1]
        if re.findall("result", file_name):
            excel_path = file_name
    
    # 担当者を抽出
    person_name = ""
    for idx, ch_list in enumerate(csv_list, 1):
        if ch_list:
            # 担当者を抽出
            with open(make_file_path_list(2, idx, 1, csv_list)) as f:

                for row, text in enumerate(f, start=0):
                    text = text.rstrip()
                    if "担当者" in text:
                        person_name = text
            person_name = re.findall('"(.*)"', person_name)[0]
            break
                
    # 担当者を追加
    df_after["Pic"] = person_name
    
    # 幅寸法データの追加
    df_after["CellWid[mm]"] = df_sample["幅寸法mm"].copy()
    
    # DB用に列を並び替え
    df_after = df_after.reindex(columns=["Type", "Klot", "LineNo", "MachineNo", "Ch", \
                                     "DischargeCap[Ah]", "ChargeRes(5sec25℃Soc50%)", "DischargeRes(10sec25℃SOC80%)", \
                                     "DischargeRes(10sec-10℃Soc30%)", "CellWid[mm]", "CellId", "Pic", "judge"])
    
    # 機種列の追加(40Aは固定)
    df_after["Type"] = "JT"
    
    # 最終的に出力するDFを二次元配列に変換
    l_2d = df_after.values.tolist()
    
    # Excelファイルの読み込み
    wb = openpyxl.load_workbook(excel_path)
    # ワークシートの読み込み
    ws = wb['試験報告書イメージ']

    # データ追加行を設定
    result_df = pd.read_excel("result.xlsx", sheet_name="試験報告書イメージ")
    result_df = result_df.where(result_df.notna(), None)
    
    max_row = len([klot for klot in result_df["Lot"] if isinstance(klot, str)])
    
    insert_row_num = max_row + 2
    write_list_2d(ws, l_2d, insert_row_num, 1)
    
    for row in ws.iter_rows(min_col=0,max_col=17,min_row=2):
        row[5].number_format='0.00000'#放電容量　小数点以下5桁
        row[6].number_format='0.000000000'#放電抵抗25-50%　小数点以下9桁
        row[7].number_format='0.000000000'#放電抵抗25-80%　小数点以下9桁
        row[8].number_format='0.000000000'#放電抵抗-10　小数点以下9桁
        row[9].number_format='0.000'#セル幅寸法　小数点以下3桁

    # Excelファイルを保存
    wb.save(excel_path)
    
    # Excelを閉じる
    wb.close()
    
##1,必要なファイルリストの作成
csv_lists = make_file_path_list()

##2,抜取セルデータ用のデータフレームを作成しDBへアップロード
sample_dfs = []
for idx, csv_list in enumerate(csv_lists):
    sample_df = make_sample_df("サンプル抜取データ.xlsx", make_file_path_list(2, csv_lists[idx][0], 1, csv_list))
    sample_dfs.append(sample_df)

for sample_df in sample_dfs:
    upload_data(insert_sample, sample_df)


##3,演算前のデータフレームを作成しDBへアップロード
before_dfs = []

for csv_list, sample_df in zip(csv_lists, sample_dfs):
    before_df = make_bofore_calc_df(csv_list[1:], sample_df)
    before_dfs.append(before_df)

for before_df in before_dfs:
    upload_data(insert_before, before_df)

##4,演算後のデータフレームを作成しDBへアップロード
after_dfs = []

for csv_list, sample_df in zip(csv_lists, sample_dfs):
    after_df = make_after_calc_df(csv_list[1:], sample_df)
    after_dfs.append(after_df)

for after_df in after_dfs:
    upload_data(insert_after, after_df)
    
##最終結果をExcelに出力
# NGおよびOKだったDFを振り分ける
ng_df_list = []
ok_df_list = []

# "judge"列(最終列)がNGの行を取り出す
for idx, after_df in enumerate(after_dfs):
    ng_data = after_df.query("judge == 'NG'")
    if len(ng_data):
        ng_df_list.append(ng_data)
    else:
        ok_df_list.append(idx)
        
if len(ng_df_list):
    # NGだったDFを一つにまとめる
    ng_df = pd.concat(ng_df_list)
    
    # 列の並び替え
    ng_df = ng_df.reindex(columns=["Date", "CellId", "Klot", "LineNo", \
                                "MachineNo", "Ch", "Approximation", "DischargeCap[Ah]", "judge_dc", \
                                "ChargeRes(5sec25℃Soc50%)", "judge_cr", 
                                "DischargeRes(10sec25℃SOC80%)", "judge_dr25", \
                                "DischargeRes(10sec-10℃Soc30%)", "judge_dr10"])
    
    # 列名を日本語に変更
    ng_df.rename(columns={'Date':'日付','CellId':'セルID', 'Klot':'検査ロット', 'LineNo':'生産ライン', 
                      'MachineNo':'投入サイクル機No', 'Ch':'チャネル', 'Approximation':'最終計算方法',
                      'DischargeCap[Ah]':'放電容量', 'judge_dc':'放電容量結果', 'ChargeRes(5sec25℃Soc50%)':'充電抵抗', 'judge_cr':'充電抵抗結果',
                      'DischargeRes(10sec25℃SOC80%)':'放電抵抗(10sec25℃SOC80%)', 'judge_dr25':'放電抵抗(10sec25℃SOC80%)結果',
                      'DischargeRes(10sec-10℃Soc30%)':'放電抵抗(10sec-10℃Soc30%)', 'judge_dr10':'放電抵抗(10sec-10℃Soc30%)結果'}, inplace=True)
    
    # 作業者にメッセージを出力
    print("NGとなった項目があります。対象箇所をExcelで出力します。再測定を実施してください。")
    
    # NGとなったDFをExcelに出力
    ng_df.to_excel('./NG項目.xlsx')


# OKだった検査ロットに関してはExcelに出力
for idx in ok_df_list:
    ch_no = [idx for idx, ch_list in enumerate(csv_lists[idx][1:]) if ch_list][0]
    output_to_excel(after_dfs[idx], sample_dfs[idx], csv_lists[idx][1:], ch_no + 1)

    
# コンソールがすぐに消えてしまわないようにするための処理
input("処理を完了するには任意のキーを押してください。")